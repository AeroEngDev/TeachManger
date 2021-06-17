import tkinter as tk
import openpyxl as pyxl
import openpyxl.chart as pyxl_chart
import os as os
import pdb
from pathlib import Path


class excel_interact:

    def __init__(self, path_name):

        self.excel_window = None

        # split path and filename:
        path: Path = Path(path_name)
        #pdb.set_trace()
        split_path = os.path.split(path_name)

        # if only a filename without a path is given,
        # stay in the cwd
        if split_path[0] != '':
            os.chdir(split_path[0])



        self.filename = split_path[1]

    def create(self, dict_data, import_graph):
        #pdb.set_trace()
        self.dict_data = dict_data
        workbook = pyxl.Workbook()
        sheet = workbook.active

        column = 1
        for heading, data_values in self.dict_data.items():
            sheet.cell(row=1, column=column, value=heading)
            for row, cell_value in enumerate(data_values):
                sheet.cell(row=row+2, column=column, value=cell_value)
            column = column + 1

        if import_graph == 1:
            img = pyxl.drawing.image.Image("temp.png")
            img.anchor = 'F1'
            sheet.add_image(img)

        workbook.save(self.filename)
        return workbook
        #return os.system(f'find {self.name}')

    def read(self):
        workbook = pyxl.load_workbook(filename=self.filename)

        # get a sheet object:
        sheet_obj = workbook.active

        # get the growth of the table:
        dimension = (sheet_obj.max_row, sheet_obj.max_column)

        # define a empty dict in which the data is stored
        read_data_list = []
        read_headings_list = []

        stud_grades_dict = {}

        # read the table columnwise:
        for k in range(1, dimension[0]+1):
            read_data_entity_list = []
            for i in range(1, dimension[1]+1):
                if k == 1:
                    stud_grades_dict[sheet_obj.cell(row=1, column=i).value] = []
                else:
                    if sheet_obj.cell(row=k, column=i).value is not None:
                        current_key_value = sheet_obj.cell(row=1, column=i).value
                        if current_key_value is None:
                            current_key_value = i

                        stud_grades_dict[current_key_value].append(sheet_obj.cell(row=k, column=i).value)
            read_data_entity_tuple = tuple(read_data_entity_list)
            if read_data_entity_tuple != ():
                read_data_list.append(read_data_entity_tuple)
        return [read_headings_list, read_data_list]

    def draw_pie_plot(self, workbook, grade_weight_data):

        worksheet = workbook.active
        get_table_size = (worksheet.max_row, worksheet.max_column)
        i = 0
        sum_list = []
        k = 0
        for key_parent, value in grade_weight_data.items():
            worksheet.cell(row=2+i, column=get_table_size[1]+5, value=key_parent)
            sum = 0
            j = 0
            for key_child, grade_value in value.items():
                worksheet.cell(column=get_table_size[1]+6+j, row=2+i, value=grade_value)
                worksheet.cell(column=get_table_size[1]+10, row=2+k, value =key_child)
                worksheet.cell(column=get_table_size[1]+11, row=2+k, value = grade_value)
                sum += grade_value
                j += 1
                k += 1
            # calculate the sum of the row and write it into the next column
            sum_list.append(sum)
            i += 1

        last_column = worksheet.max_column+1
        for pos_sum, sum in enumerate(sum_list):
            worksheet.cell(column=last_column, row=pos_sum+2, value=sum)

        pie = pyxl_chart.DoughnutChart()
        label = pyxl_chart.Reference(worksheet, min_col=get_table_size[1]+10, min_row=2, max_row=2+k-1)
        data = pyxl_chart.Reference(worksheet, min_col=get_table_size[1]+11, min_row=1, max_row=2+k-1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(label)
        pie.title = "Notengewichtung"
        # put the grade_child values as projected pies
        outer_pie_labels = pyxl_chart.Reference(worksheet, min_row=2, max_row=2+i-1, min_col=get_table_size[1]+5)
        outer_pie_data = pyxl_chart.Reference(worksheet, min_col=last_column, min_row=1, max_row=2+pos_sum)
        series2 = pyxl_chart.Series(outer_pie_data, title_from_data=True)
        pie.series.append(series2)

        worksheet.add_chart(pie, "G7")

        workbook.save(self.name)

# '    def import_pie_chart_image(self, PieChart_obj, filename):
#         """ Methods calls Pie Chart to export the Canvas as a .png image
#         """
#
#         # call PiePlot export method:
#         PieChart_obj.export_piechart_as_gif(filename)
#         self.filename_graph = filename + ".png"'

    def get_grade_weight_plot(self):
        # get pie out of self.name
        pass
